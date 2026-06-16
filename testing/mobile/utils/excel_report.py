import openpyxl
from openpyxl.styles import PatternFill
import os
import datetime

def create_excel_report(test_results, filename='report_mobile.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Results"

    # Define headers
    headers = ['Test ID', 'Category', 'Test Case Name', 'Status', 'Error/Details', 'Timestamp']
    ws.append(headers)

    # Styles
    pass_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    fail_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

    for result in test_results:
        row = [
            result.get('id'),
            result.get('category'),
            result.get('name'),
            result.get('status'),
            result.get('details'),
            result.get('timestamp')
        ]
        ws.append(row)
        
        # Color coding
        last_row = ws.max_row
        status_cell = ws.cell(row=last_row, column=4) # Status is 4th col
        if result.get('status') == 'Pass':
            status_cell.fill = pass_fill
        else:
            status_cell.fill = fail_fill

    # Save
    save_path = os.path.join(os.path.dirname(__file__), '..', '..', filename)
    wb.save(save_path)
    print(f"Mobile Report generated: {save_path}")
